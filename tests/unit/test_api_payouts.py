"""The payout HTTP surface — the wall, the gates, and the money on the wire.

Four things are worth pinning about `app/api/payouts.py` and all four are here:

* **R5, the commercials wall.** We are on a `BYPASSRLS` connection, so nothing in
  the database refuses an LDE Executive a trainer's net pay. The refusal has to be
  in the router, and it has to happen before any row is read — so the fake session
  below records every query and the wall tests assert the count is zero.
* **R7, no float.** Both directions: a JSON float is a 422 on the way in, and no
  amount is ever a JSON number on the way out. The outbound test parses the
  response with `parse_float` wired to fail, so it catches an amount anywhere in
  the body rather than only in the fields somebody remembered to assert.
* **R2 through the API.** Both CLAUDE.md §6 regression fixtures are driven end to
  end over HTTP and must still reconcile to the rupee. If the router ever grows
  its own arithmetic, these break.
* **R4.** No route this service exposes can release anything.

The session is faked rather than mocked per call: it holds ORM instances and
answers `get()` by primary key and `execute()` by the selected entity. It
deliberately does NOT evaluate WHERE clauses — a test puts exactly the rows the
scenario has in it — and it raises on `add`/`commit`, which is how "these
endpoints write nothing" is asserted for free on every test in the file.
"""

from __future__ import annotations

import datetime as dt
import io
import json
import uuid
from collections.abc import Iterator
from decimal import Decimal
from typing import Any

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api import payouts
from app.core.audit import AuditEvent, AuditWriter, get_audit_writer
from app.core.security import Principal, get_principal
from app.db.models import (
    Batch,
    College,
    Deployment,
    Profile,
    Program,
    RemunerationSheet,
    Trainer,
    TrainerAttendance,
    TrainerBankAccount,
    WorkOrder,
)
from app.db.session import get_session
from app.domain.enums import (
    ArtifactState,
    AttendanceMark,
    DocStatus,
    Persona,
    ProgramType,
    RateBasis,
    ValidationCode,
)
from app.services.remuneration.generators import INVOICE_COLUMNS, remuneration_columns

# --- identity ----------------------------------------------------------------

COLLEGE_ID = uuid.UUID("11111111-1111-1111-1111-111111111111")
PROGRAM_ID = uuid.UUID("22222222-2222-2222-2222-222222222222")
BATCH_ID = uuid.UUID("33333333-3333-3333-3333-333333333333")
DEPLOYMENT_ID = uuid.UUID("44444444-4444-4444-4444-444444444444")
TRAINER_ID = uuid.UUID("55555555-5555-5555-5555-555555555555")
OTHER_TRAINER_ID = uuid.UUID("66666666-6666-6666-6666-666666666666")
STAFF_USER_ID = uuid.UUID("77777777-7777-7777-7777-777777777777")
TRAINER_USER_ID = uuid.UUID("88888888-8888-8888-8888-888888888888")

#: CLAUDE.md §6 fixture one: VEMA PRUDHVI SAI, bCAP, 80,000/month, 26-31 Jul 2026,
#: TA&DA 100 -> Earned 15,484 · Gross 15,584 · TDS 1,548 · Net 14,035.
PAN = "VEMAP1234K"
PERIOD_START = dt.date(2026, 7, 26)
PERIOD_END = dt.date(2026, 7, 31)

ENDPOINTS = (
    "/payouts/preview",
    "/payouts/validate",
    "/payouts/remuneration-sheet.xlsx",
    "/payouts/invoice-sheet.xlsx",
)

#: Every endpoint except the preview: the three a trainer must never reach.
INTERNAL_ONLY_ENDPOINTS = ENDPOINTS[1:]


# --- the fake session ---------------------------------------------------------


class FakeResult:
    """Just enough of `sqlalchemy.Result` for the four call shapes this router uses."""

    def __init__(self, rows: list[Any]) -> None:
        self._rows = rows

    def scalar_one_or_none(self) -> Any:
        if not self._rows:
            return None
        assert len(self._rows) == 1, "scalar_one_or_none() over more than one row"
        return self._rows[0]

    def scalars(self) -> FakeResult:
        return self

    def all(self) -> list[Any]:
        return list(self._rows)


class FakeSession:
    """An in-memory stand-in for `AsyncSession`, keyed by mapped class.

    WHERE clauses are not evaluated: a test loads exactly the rows its scenario
    has, so a filter bug would show up as wrong data rather than as a silently
    empty result. `get()` is real, though — it matches on primary key, which is
    what the authorisation path depends on.

    Writing raises. These endpoints compute, validate and generate; nothing here
    may persist or transition anything (R4), and that holds for every test in the
    file without any of them asserting it.
    """

    def __init__(self, *rows: Any) -> None:
        self.rows = list(rows)
        self.queries: list[str] = []

    def _of(self, model: type[Any]) -> list[Any]:
        return [row for row in self.rows if isinstance(row, model)]

    async def get(self, model: type[Any], pk: Any) -> Any:
        self.queries.append(model.__name__)
        return next((row for row in self._of(model) if row.id == pk), None)

    async def execute(self, statement: Any) -> FakeResult:
        entity = statement.column_descriptions[0]["entity"]
        self.queries.append(entity.__name__)
        return FakeResult(self._of(entity))

    def add(self, _obj: Any) -> None:  # pragma: no cover - must never be called
        raise AssertionError("a payout endpoint tried to write (R4)")

    def add_all(self, _objs: Any) -> None:  # pragma: no cover
        raise AssertionError("a payout endpoint tried to write (R4)")

    async def commit(self) -> None:  # pragma: no cover
        raise AssertionError("a payout endpoint tried to commit (R4)")


class RecordingAudit(AuditWriter):
    """An `AuditWriter` that keeps what it was given instead of logging it."""

    def __init__(self) -> None:
        self.events: list[AuditEvent] = []

    async def write(self, event: AuditEvent) -> None:
        self.events.append(event)


# --- scenario -----------------------------------------------------------------


def scenario(
    *,
    program_type: ProgramType = ProgramType.BCAP,
    marks: dict[dt.date, AttendanceMark] | None = None,
    work_order: WorkOrder | None | str = "default",
    trainer_pan: str = PAN,
    sheets: list[RemunerationSheet] | None = None,
    deployed_trainer_id: uuid.UUID = TRAINER_ID,
    trainer_login_owns: uuid.UUID | None = TRAINER_ID,
    bank: TrainerBankAccount | None | str = None,
) -> FakeSession:
    """One college, one program, one batch, one deployed trainer, marked present.

    `work_order="default"` installs a signed 80,000/month order covering the whole
    fiscal year; pass `None` for a trainer with no order on file, or a `WorkOrder`
    to vary status, rate or validity window.

    `deployed_trainer_id` and `trainer_login_owns` are the two halves of the
    ownership check — who the deployment is for, and which `trainers` row the
    trainer login is attached to through `profiles.trainer_id`. Setting them apart
    is how "somebody else's payout" is built.

    `bank` defaults to NO rails on file, because that is the state every trainer
    starts in and the state §7 must block; pass `"default"` for a well-formed row
    from `trainer_bank_accounts` (1400), or a `TrainerBankAccount` to vary it.
    """
    period_marks = marks if marks is not None else _all_present(PERIOD_START, PERIOD_END)

    order = _work_order() if work_order == "default" else work_order
    rows: list[Any] = [
        College(id=COLLEGE_ID, name="Malineni Lakshmaiah"),
        Program(id=PROGRAM_ID, college_id=COLLEGE_ID, type=program_type, name="bCAP 2026"),
        Batch(id=BATCH_ID, program_id=PROGRAM_ID, name="CSE-A"),
        Deployment(id=DEPLOYMENT_ID, trainer_id=deployed_trainer_id, batch_id=BATCH_ID),
        Trainer(
            id=deployed_trainer_id,
            pan=trainer_pan,
            full_name="VEMA PRUDHVI SAI",
            email="vema@example.com",
            phone="9000000000",
            zoho_id="ZOHO-1",
        ),
        Profile(
            id=TRAINER_USER_ID,
            role=Persona.TRAINER,
            is_admin=False,
            trainer_id=trainer_login_owns,
        ),
    ]
    rows += [
        TrainerAttendance(id=uuid.uuid4(), deployment_id=DEPLOYMENT_ID, mark_date=day, mark=mark)
        for day, mark in sorted(period_marks.items())
    ]
    if order is not None:
        rows.append(order)
    rails = _bank_account(trainer_id=deployed_trainer_id) if bank == "default" else bank
    if rails is not None:
        rows.append(rails)
    rows += sheets or []
    return FakeSession(*rows)


def _bank_account(*, trainer_id: uuid.UUID = TRAINER_ID, **overrides: Any) -> TrainerBankAccount:
    """Well-formed rails: 14 digits and an RBI-shaped IFSC, both §7-clean."""
    spec: dict[str, Any] = {
        "trainer_id": trainer_id,
        "bank_account_number": "50100123456789",
        "ifsc": "HDFC0001234",
        "bank_name": "HDFC Bank",
        "branch": "Guntur",
        "account_name": "VEMA PRUDHVI SAI",
    }
    spec.update(overrides)
    return TrainerBankAccount(**spec)


def _all_present(start: dt.date, end: dt.date) -> dict[dt.date, AttendanceMark]:
    days = (end - start).days + 1
    return {start + dt.timedelta(days=n): AttendanceMark.PRESENT for n in range(days)}


def _work_order(**overrides: Any) -> WorkOrder:
    spec: dict[str, Any] = {
        "id": uuid.uuid4(),
        "trainer_id": TRAINER_ID,
        "program_id": PROGRAM_ID,
        "rate": Decimal("80000"),
        "rate_basis": RateBasis.PER_MONTH,
        "valid_from": dt.date(2026, 4, 1),
        "valid_to": dt.date(2027, 3, 31),
        "status": DocStatus.SIGNED,
    }
    spec.update(overrides)
    return WorkOrder(**spec)


def principal(
    persona: Persona, *, reach: bool = True, user_id: uuid.UUID | None = None
) -> Principal:
    return Principal(
        user_id=user_id or (TRAINER_USER_ID if persona is Persona.TRAINER else STAFF_USER_ID),
        persona=persona,
        college_ids=frozenset({COLLEGE_ID}) if reach else frozenset(),
    )


@pytest.fixture
def audit() -> RecordingAudit:
    return RecordingAudit()


@pytest.fixture
def client(audit: RecordingAudit) -> Iterator[TestClient]:
    """A one-router app.

    `app.main.create_app()` is not used: it builds `Settings` from the
    environment, which would make this suite depend on a developer's `.env`.
    Registration in `main.py` has its own test at the bottom of the file.
    """
    app = FastAPI()
    app.include_router(payouts.router)
    app.dependency_overrides[get_audit_writer] = lambda: audit
    with TestClient(app) as test_client:
        yield test_client


def call(
    client: TestClient,
    session: FakeSession,
    caller: Principal,
    url: str = "/payouts/preview",
    **body: Any,
) -> Any:
    """POST `url` as `caller`, against `session`."""
    client.app.dependency_overrides[get_session] = lambda: session
    client.app.dependency_overrides[get_principal] = lambda: caller
    payload: dict[str, Any] = {
        "deployment_id": str(DEPLOYMENT_ID),
        "period_start": PERIOD_START.isoformat(),
        "period_end": PERIOD_END.isoformat(),
    }
    payload.update(body)
    return client.post(url, json=payload)


# =============================================================================
# R5 — the commercials wall
# =============================================================================


@pytest.mark.parametrize("url", ENDPOINTS)
@pytest.mark.parametrize("persona", [Persona.LDE_EXECUTIVE, Persona.COLLEGE])
def test_a_persona_outside_the_commercials_wall_gets_nothing(client, url, persona) -> None:
    """CLAUDE.md §4: an LDE Executive has NO commercials. A payout is a commercial.

    "Zero rows" is asserted literally: the refusal lands before a single query, so
    the endpoint cannot even confirm the deployment exists. On a `BYPASSRLS`
    connection that ordering is the whole protection — a query issued first would
    already have returned the rate.
    """
    session = scenario()
    response = call(client, session, principal(persona), url=url)

    assert response.status_code == 403
    assert session.queries == []


@pytest.mark.parametrize("url", ENDPOINTS)
def test_a_manager_without_reach_is_refused(client, url) -> None:
    """The wall and the scope are separate conjuncts (0700_finance.sql).

    A Manager satisfies `can_see_commercials()` and still may not price a trainer
    in a college they are not assigned to. Dropping `require_college_reach()`
    would leave a passing wall test and a cross-cluster leak.
    """
    session = scenario()
    response = call(client, session, principal(Persona.MANAGER, reach=False), url=url)

    assert response.status_code == 403
    assert "college" in response.json()["detail"].lower()


@pytest.mark.parametrize("persona", [Persona.MANAGER, Persona.SENIOR_MANAGER])
def test_the_two_commercial_personas_may_preview(client, persona) -> None:
    assert call(client, scenario(), principal(persona)).status_code == 200


# =============================================================================
# R5 — a trainer sees NOTHING here
# =============================================================================
# This section once read "a trainer sees their own payout and nothing else" and
# asserted a 200 for a trainer previewing their own figures. That was true of
# the three-persona build. It is not true now.
#
# §4 (owner's decision, 2026-08-18): trainers are RECORDS, not users. Migration
# 1800 dropped all eighteen trainer policies, and SEC-06 removed the last
# `trainer_may_read` carve-out from `_require_payout_persona`, so
# `/payouts/preview` is behind `require_commercials()` alone — Senior Manager and
# Manager.
#
# The two tests deleted here asserted the carve-out itself, so there was no way
# to keep them and close the finding: one expected 200 for a trainer reading
# their own payout, the other expected the old "You do not have access to this
# payout" detail string. What survives is the deny-by-default case below, which
# is the half that still means something.


def test_a_trainer_with_no_trainer_record_reaches_nothing(client) -> None:
    """A trainer login not linked to a `trainers` row is deny-by-default.

    Defaulting to "no link, no check" is how an unfinished onboarding becomes an
    open door.
    """
    session = scenario(trainer_login_owns=None)
    assert call(client, session, principal(Persona.TRAINER)).status_code == 403


def test_a_trainer_learns_nothing_from_probing_deployment_ids(client) -> None:
    """A missing deployment and someone else's deployment must look identical.

    A 404 here would make the endpoint an id oracle: a trainer could walk the
    deployment table and learn its size and shape. Internal callers, who have
    already cleared persona and reach, do get the 404 — see the next test.
    """
    # The `own` control that used to sit here asserted a 200 and is gone with the
    # carve-out (SEC-06). The oracle property it guarded is unchanged and is
    # arguably stronger now: every deployment looks the same to a trainer,
    # including one that is genuinely theirs.
    unknown = call(client, scenario(), principal(Persona.TRAINER), deployment_id=str(uuid.uuid4()))
    not_mine = call(
        client, scenario(deployed_trainer_id=OTHER_TRAINER_ID), principal(Persona.TRAINER)
    )
    own = call(client, scenario(), principal(Persona.TRAINER))

    assert unknown.status_code == not_mine.status_code == own.status_code == 403
    assert unknown.json() == not_mine.json() == own.json(), (
        "a trainer must not be able to tell their own deployment from a stranger's "
        "or from one that does not exist"
    )


def test_an_internal_caller_gets_404_for_a_deployment_that_does_not_exist(client) -> None:
    response = call(client, scenario(), principal(Persona.MANAGER), deployment_id=str(uuid.uuid4()))
    assert response.status_code == 404


@pytest.mark.parametrize("url", INTERNAL_ONLY_ENDPOINTS)
def test_a_trainer_is_refused_the_report_and_both_sheets(client, url) -> None:
    """§4 ends "Nothing else."

    The §7 report carries the work-order rate and the trailing net-pay average,
    and the sheets are Finance artifacts. A trainer's own preview is the whole of
    their access.
    """
    session = scenario()
    response = call(client, session, principal(Persona.TRAINER), url=url)

    assert response.status_code == 403
    assert session.queries == []


# =============================================================================
# R2 / R6 — the §6 fixtures, driven over HTTP
# =============================================================================


def test_the_vema_fixture_reconciles_to_the_rupee_through_the_api(client) -> None:
    """CLAUDE.md §6: Earned 15,484 · Gross 15,584 · TDS 1,548 · **Net 14,035**.

    Driven end to end: attendance is counted from marks in the database, the rate
    comes off the signed work order, and only `engine.py` multiplies anything. If
    the router ever grows its own arithmetic, this is what catches it.
    """
    response = call(client, scenario(), principal(Persona.MANAGER), ta_da="100")
    breakdown = response.json()["breakdown"]

    assert response.status_code == 200
    assert breakdown["net"] == "14035"
    assert _paise(breakdown["earned"]) == Decimal("15483.87")
    assert _paise(breakdown["gross"]) == Decimal("15583.87")
    assert _paise(breakdown["tds"]) == Decimal("1548.39")
    assert breakdown["net_in_words"] == "Fourteen Thousand and Thirty Five Rupees Only"
    assert breakdown["rate_source"] == "work_order"


def test_the_bushily_fixture_reconciles_to_the_rupee_through_the_api(client) -> None:
    """CLAUDE.md §6: 65,000/month, full July 2026 -> Earned 65,000, **Net 58,500**.

    The multiply-before-divide case. `65000 / 31 * 31` is 64999.99999…, and a
    full month must recombine exactly — over HTTP as much as in the engine.
    """
    july = (dt.date(2026, 7, 1), dt.date(2026, 7, 31))
    session = scenario(
        marks=_all_present(*july),
        work_order=_work_order(rate=Decimal("65000")),
    )
    response = call(
        client,
        session,
        principal(Persona.MANAGER),
        period_start=july[0].isoformat(),
        period_end=july[1].isoformat(),
    )
    breakdown = response.json()["breakdown"]

    assert breakdown["payable_days"] == "31"
    assert _paise(breakdown["earned"]) == Decimal("65000.00")
    assert _paise(breakdown["tds"]) == Decimal("6500.00")
    assert breakdown["net"] == "58500"


def test_rate_per_day_is_reported_but_does_not_reproduce_earned(client) -> None:
    """R6, made visible on the wire.

    `rate_per_day` is the sheet's display column. On the per-month path it is a
    repeating decimal, and multiplying it back out is the legacy spreadsheet's
    bug — 2581 x 6 = 15,486 against a correct 15,483.87. The API must expose the
    display value without it ever having touched the money.
    """
    breakdown = call(client, scenario(), principal(Persona.MANAGER)).json()["breakdown"]

    assert _paise(breakdown["rate_per_day"]) == Decimal("2580.65")
    assert Decimal(breakdown["rate_per_day"]) * Decimal(breakdown["payable_days"]) != Decimal(
        breakdown["earned"]
    )


# =============================================================================
# R7 — no float, in either direction
# =============================================================================


@pytest.mark.parametrize("field", ["ta_da", "accommodation", "travel_reimbursement", "deductions"])
def test_a_json_float_amount_is_refused(client, field) -> None:
    """Pydantic would coerce 100.5 into Decimal("100.5") without complaint.

    R7 has no exceptions, and `app.domain.money.money()` refuses a float rather
    than converting it so the defect surfaces where it is cheap to find. The API
    boundary runs the same guard, so an amount is a string or an integer on the
    wire and never a JSON float.
    """
    response = call(client, scenario(), principal(Persona.MANAGER), **{field: 100.50})

    assert response.status_code == 422
    assert "float" in response.text.lower()


@pytest.mark.parametrize("value", ["100.50", 100])
def test_a_string_or_integer_amount_is_accepted(client, value) -> None:
    assert call(client, scenario(), principal(Persona.MANAGER), ta_da=value).status_code == 200


def test_a_float_rate_override_is_refused(client) -> None:
    """The rate is the largest number in the chain; it is also the easiest to fat-finger."""
    response = call(
        client,
        scenario(),
        principal(Persona.MANAGER),
        rate=80000.5,
        rate_basis=RateBasis.PER_MONTH.value,
    )
    assert response.status_code == 422


@pytest.mark.parametrize("url", ["/payouts/preview", "/payouts/validate"])
def test_no_amount_is_ever_serialised_as_a_json_float(client, url) -> None:
    """The outbound half of R7, asserted over the whole body rather than field by field.

    `parse_float` fires on any JSON number with a decimal point anywhere in the
    response, so an amount added later that forgets the `str` serialiser fails
    here without anyone remembering to extend the test.
    """

    def refuse(raw: str) -> float:  # pragma: no cover - the assertion is that it never runs
        raise AssertionError(f"a JSON float {raw!r} reached the client — R7")

    response = call(client, scenario(), principal(Persona.MANAGER), ta_da="100.55", url=url)

    assert response.status_code == 200
    json.loads(response.text, parse_float=refuse)
    assert '"net":"14036"' in response.text.replace(" ", "")


def _paise(value: str) -> Decimal:
    """Quantise a wire amount for comparison. Display-side only (R6)."""
    return Decimal(value).quantize(Decimal("0.01"))


# =============================================================================
# §7 — the gates are surfaced, not swallowed
# =============================================================================


def test_a_blocked_report_comes_back_as_200_with_its_reasons(client) -> None:
    """A blocked payout is an answer, not an error.

    Returning 4xx would hand a Manager an error page instead of the list of things
    to fix, and every gate runs, so the whole list arrives at once rather than one
    blocker per round trip.
    """
    response = call(client, scenario(), principal(Persona.MANAGER), url="/payouts/validate")
    report = response.json()["report"]

    assert response.status_code == 200
    assert report["is_blocked"] is True
    assert report["can_submit"] is False
    # No `trainer_bank_accounts` row for this trainer, so §7 blocks and says why
    # rather than the router inventing a payment instruction.
    codes = {issue["code"] for issue in report["blocking"]}
    assert ValidationCode.BANK_ACCOUNT_MISSING.value in codes
    assert ValidationCode.IFSC_INVALID.value in codes


# =============================================================================
# §7 — the bank rails (1400), which is what makes `can_submit` reachable at all
# =============================================================================


def test_rails_on_file_clear_both_bank_gates_and_can_submit_becomes_true(client) -> None:
    """The whole point of migration 1400.

    Before it, `bank_account_missing` and `ifsc_invalid` fired on every payout in
    the system, so `can_submit` was false for every cycle and none could ever
    reach PENDING_APPROVAL. With rails on file and nothing else wrong, the §7
    verdict is clean and the payout may leave DRAFT.
    """
    report = _report(client, scenario(bank="default"))

    assert _codes(report["blocking"]) == set()
    assert report["is_blocked"] is False
    assert report["can_submit"] is True


def test_rails_are_read_from_the_database_and_never_from_the_request(client) -> None:
    """R1: payment rails are structured input from a system of record.

    `PayoutRequest` is `extra="forbid"`, so a caller trying to supply their own
    account number is a 422 rather than a payout quietly pointed somewhere new.
    """
    response = call(
        client,
        scenario(bank="default"),
        principal(Persona.MANAGER),
        url="/payouts/validate",
        bank_account_number="99999999999",
    )
    assert response.status_code == 422


def test_a_malformed_ifsc_on_file_still_blocks(client) -> None:
    """The database checks length and case; the RBI shape is §7's to check.

    `HDFC1001234` is 11 uppercase characters and would satisfy every CHECK
    constraint in 1400 — the reserved zero in position five is missing, which
    fails at the bank after release unless `gate_ifsc` catches it here.
    """
    report = _report(client, scenario(bank=_bank_account(ifsc="HDFC1001234")))

    assert ValidationCode.IFSC_INVALID.value in _codes(report["blocking"])
    assert report["can_submit"] is False


def test_the_sheets_print_the_rails_that_are_on_file(client) -> None:
    """§11: the sheet is an output contract, and the rail cells are part of it."""
    response = call(
        client,
        scenario(bank="default"),
        principal(Persona.MANAGER),
        url="/payouts/invoice-sheet.xlsx",
    )
    row = _sheet_row(response)

    # A STRING, not a number. An account number with a leading zero that Excel
    # helpfully renders as an integer is a payment that bounces.
    assert row["Bank AC no."] == "50100123456789"
    assert row["IFSC"] == "HDFC0001234"
    assert row["Name of Bank"] == "HDFC Bank"
    assert row["Branch"] == "Guntur"
    assert row["account_name"] == "VEMA PRUDHVI SAI"


def test_the_rail_cells_stay_empty_when_no_rails_are_on_file(client) -> None:
    """`generators.py`: an un-releasable payout should "look obviously incomplete
    rather than plausible". A blank is honest; an invented account number is not."""
    response = call(
        client, scenario(), principal(Persona.MANAGER), url="/payouts/invoice-sheet.xlsx"
    )
    row = _sheet_row(response)

    assert row["Bank AC no."] is None
    assert row["IFSC"] is None
    assert row["account_name"] is None


def test_the_preview_still_computes_for_a_blocked_payout(client) -> None:
    """Compute and permission are separate questions, and §7 needs the answer to the
    first to ask the second — `net > 0` is itself a gate."""
    response = call(client, scenario(), principal(Persona.MANAGER), ta_da="100")
    assert response.status_code == 200
    assert response.json()["breakdown"]["net"] == "14035"


def test_an_unsigned_work_order_blocks(client) -> None:
    """§7: a signed work order must be on file. The rate is still readable off an
    unsigned order, so the payout computes and the gate is what stops it."""
    session = scenario(work_order=_work_order(status=DocStatus.SENT))
    report = _report(client, session)

    assert ValidationCode.WORK_ORDER_MISSING.value in _codes(report["blocking"])
    assert report["is_blocked"] is True


def test_a_period_outside_the_work_order_window_blocks(client) -> None:
    session = scenario(work_order=_work_order(valid_to=dt.date(2026, 7, 20)))
    assert ValidationCode.WORK_ORDER_PERIOD_MISMATCH.value in _codes(
        _report(client, session)["blocking"]
    )


def test_a_rate_override_that_disagrees_with_the_work_order_blocks(client) -> None:
    """A caller-supplied rate is an assertion to be checked, never an override.

    §7: "Engagement rate matches the rate in the signed WO." The work order is the
    contract; if the request disagrees with it, one of the two is wrong and
    neither the engine nor a Manager should guess which.
    """
    report = _report(
        client,
        scenario(),
        rate="90000",
        rate_basis=RateBasis.PER_MONTH.value,
    )
    assert ValidationCode.RATE_MISMATCH_WITH_WORK_ORDER.value in _codes(report["blocking"])


def test_incomplete_attendance_blocks_on_crt_and_only_warns_on_bcap(client) -> None:
    """CLAUDE.md §5's asymmetry, carried intact through the API.

    An unmarked day silently PAYS a bCAP trainer and silently UNDERPAYS a CRT one,
    so the same condition is a hard block on one path and a stated-reason warning
    on the other. A router that flattened the two severities would ship silent
    underpayments.
    """
    partial = {PERIOD_START: AttendanceMark.PRESENT}

    bcap = _report(client, scenario(marks=partial))
    crt = _report(
        client,
        scenario(
            program_type=ProgramType.CRT,
            marks=partial,
            work_order=_work_order(rate=Decimal("3500"), rate_basis=RateBasis.PER_DAY),
        ),
    )

    assert ValidationCode.ATTENDANCE_INCOMPLETE.value in _codes(bcap["warnings"])
    assert ValidationCode.ATTENDANCE_INCOMPLETE.value in _codes(crt["blocking"])


def test_a_warning_needs_a_stated_reason(client) -> None:
    """§7 warnings are "permitted, but require a stated reason"."""
    partial = {PERIOD_START: AttendanceMark.PRESENT}
    code = ValidationCode.ATTENDANCE_INCOMPLETE.value

    silent = _report(client, scenario(marks=partial))
    explained = _report(
        client,
        scenario(marks=partial),
        stated_reasons={code: "Weekend marks pending from the LDE Executive"},
    )
    blank = _report(client, scenario(marks=partial), stated_reasons={code: "   "})

    assert code in silent["reasons_missing"]
    assert code not in explained["reasons_missing"]
    assert (
        code in blank["reasons_missing"]
    ), "a blank reason is how a required field becomes a formality"


def test_a_malformed_pan_reports_the_gate_instead_of_500ing(client) -> None:
    """The invoice number is seeded from PAN and `build_invoice_number()` raises on
    a bad one. The caller asked what this payout looks like; the honest answer is
    the full breakdown plus the gate that explains the missing number."""
    response = call(client, scenario(trainer_pan="NOTAPAN"), principal(Persona.MANAGER))

    assert response.status_code == 200
    assert response.json()["invoice_number"] is None

    report = _report(client, scenario(trainer_pan="NOTAPAN"))
    assert ValidationCode.PAN_INVALID.value in _codes(report["blocking"])


def test_a_previously_issued_invoice_number_advances_the_sequence(client) -> None:
    """§6: uniqueness is on (pan, fiscal_year, month, seq), and the fiscal year
    derives from the payout month — July 2026 is 26-27."""
    session = scenario(
        sheets=[
            RemunerationSheet(
                id=uuid.uuid4(),
                trainer_id=TRAINER_ID,
                program_id=PROGRAM_ID,
                period_start=dt.date(2026, 7, 1),
                period_end=dt.date(2026, 7, 25),
                invoice_fy="26-27",
                invoice_month="JUL",
                invoice_seq=1,
                invoice_no="VEMA/26-27/JUL1",
                net_amount=Decimal("50000"),
            )
        ]
    )
    body = call(client, session, principal(Persona.MANAGER)).json()

    assert body["invoice_number"] == "VEMA/26-27/JUL2"


def _report(client: TestClient, session: FakeSession, **body: Any) -> Any:
    response = call(client, session, principal(Persona.MANAGER), url="/payouts/validate", **body)
    assert response.status_code == 200, response.text
    return response.json()["report"]


def _codes(issues: list[dict[str, Any]]) -> set[str]:
    return {issue["code"] for issue in issues}


def _sheet_row(response: Any) -> dict[Any, Any]:
    """The single data row of a downloaded workbook, keyed by its header cell."""
    sheet = load_workbook(io.BytesIO(response.content)).active
    header = [cell.value for cell in sheet[1]]
    return {name: sheet.cell(row=2, column=i).value for i, name in enumerate(header, start=1)}


# =============================================================================
# Request validation
# =============================================================================


def test_a_period_spanning_two_months_is_refused(client) -> None:
    """`days_in_month` prorates the retainer, the sheet's eighth header is bound to
    the month and the invoice fiscal year derives from it. A straddling period has
    no single right answer for any of the three."""
    response = call(
        client,
        scenario(),
        principal(Persona.MANAGER),
        period_start="2026-07-26",
        period_end="2026-08-05",
    )
    assert response.status_code == 422
    assert "two months" in response.text


def test_a_rate_without_a_basis_is_refused(client) -> None:
    """§5 puts the basis on the work order; inferring it from the program type
    would be the router inventing a contract term."""
    response = call(client, scenario(), principal(Persona.MANAGER), rate="80000")
    assert response.status_code == 422


def test_no_work_order_and_no_rate_is_unprocessable(client) -> None:
    response = call(client, scenario(work_order=None), principal(Persona.MANAGER))
    assert response.status_code == 422
    assert "rate" in response.text.lower()


def test_a_rate_supplied_without_a_work_order_computes_and_says_so(client) -> None:
    body = call(
        client,
        scenario(work_order=None),
        principal(Persona.MANAGER),
        rate="80000",
        rate_basis=RateBasis.PER_MONTH.value,
    ).json()

    assert body["breakdown"]["rate_source"] == "request_override"
    assert body["breakdown"]["net"] == "13935"


def test_an_unknown_field_is_refused_rather_than_ignored(client) -> None:
    """`extra="forbid"`: a mistyped `ta_da` must be a 422, not a silent zero."""
    response = call(client, scenario(), principal(Persona.MANAGER), taa_da="100")
    assert response.status_code == 422


def test_a_duplicated_attendance_day_is_reported_not_collapsed(client) -> None:
    """§6 is one row per day. A duplicated `A` deducts twice on the bCAP path, and
    a dict comprehension would silently keep whichever row came last."""
    session = scenario()
    session.rows.append(
        TrainerAttendance(
            id=uuid.uuid4(),
            deployment_id=DEPLOYMENT_ID,
            mark_date=PERIOD_START,
            mark=AttendanceMark.ABSENT,
        )
    )
    response = call(client, session, principal(Persona.MANAGER))

    assert response.status_code == 409
    assert str(PERIOD_START) in response.json()["detail"]


# =============================================================================
# The sheets — R4 and the §11 column contract
# =============================================================================


def test_the_remuneration_sheet_keeps_the_legacy_21_columns(client) -> None:
    """§11: "Sheet outputs preserve legacy column order. People trust the format
    they read." Including the misspelled `Accomodation`, which Finance's VLOOKUPs
    are written against."""
    response = call(
        client,
        scenario(),
        principal(Persona.MANAGER),
        url="/payouts/remuneration-sheet.xlsx",
        ta_da="100",
    )
    sheet = load_workbook(io.BytesIO(response.content)).active

    assert response.status_code == 200
    assert response.headers["content-type"] == payouts.XLSX_MEDIA_TYPE
    assert [cell.value for cell in sheet[1]] == list(remuneration_columns(PERIOD_START))
    assert "Accomodation" in [cell.value for cell in sheet[1]]
    assert sheet.cell(row=2, column=17).value == 14035  # Net Pay


def test_the_invoice_sheet_keeps_the_legacy_34_columns_and_words_the_amount(client) -> None:
    """§6: the legacy sheet renders `#NAME?` from a missing Excel macro. Not reproduced."""
    response = call(
        client,
        scenario(),
        principal(Persona.MANAGER),
        url="/payouts/invoice-sheet.xlsx",
        ta_da="100",
    )
    sheet = load_workbook(io.BytesIO(response.content)).active
    header = [cell.value for cell in sheet[1]]
    row = {name: sheet.cell(row=2, column=i).value for i, name in enumerate(header, start=1)}

    assert header == list(INVOICE_COLUMNS)
    assert row["Amount in Words"] == "Fourteen Thousand and Thirty Five Rupees Only"
    assert row["Net Pay"] == row["Total Pay"] == 14035
    assert row["Expense for the month"] is None, "§14 Q2 is open — do not invent a value"
    assert row["AM Mail ID"] is None, "findings §1 — the recipient is unidentified"


@pytest.mark.parametrize("url", ["/payouts/remuneration-sheet.xlsx", "/payouts/invoice-sheet.xlsx"])
def test_a_generated_sheet_is_labelled_a_draft_and_carries_its_verdict(client, url) -> None:
    """R4: nothing leaves this system unapproved.

    The sheet generates even though §7 blocks — this scenario has no rails on
    file, and `generators.py` wants an un-releasable payout to look obviously
    incomplete rather than plausible. What stops it being mistaken for a
    Finance-ready document is the labelling, not the refusal.
    """
    response = call(client, scenario(), principal(Persona.MANAGER), url=url)

    assert response.headers[payouts.ARTIFACT_STATE_HEADER] == ArtifactState.DRAFT.value
    assert response.headers[payouts.BLOCKED_HEADER] == "true"
    assert (
        ValidationCode.BANK_ACCOUNT_MISSING.value in response.headers[payouts.BLOCKING_CODES_HEADER]
    )
    assert "DRAFT" in response.headers["content-disposition"]


def test_a_sheet_is_refused_when_the_pan_cannot_seed_an_invoice_number(client) -> None:
    """A blank invoice number in a document Finance files is worse than no document."""
    response = call(
        client,
        scenario(trainer_pan="NOTAPAN"),
        principal(Persona.MANAGER),
        url="/payouts/invoice-sheet.xlsx",
    )
    assert response.status_code == 409
    assert "/payouts/validate" in response.json()["detail"]


# =============================================================================
# §11 — audit, and R3/R4 — no release capability
# =============================================================================


@pytest.mark.parametrize(
    ("url", "action"),
    [
        ("/payouts/validate", "payout.validated"),
        ("/payouts/remuneration-sheet.xlsx", "payout.remuneration_sheet_generated"),
        ("/payouts/invoice-sheet.xlsx", "payout.invoice_sheet_generated"),
    ],
)
def test_every_verdict_and_artifact_writes_an_audit_event(client, audit, url, action) -> None:
    """§11: actor, action, before, after, at — and no float in the `after` snapshot.

    An audit row that rounds the figure it is attesting to is worse than no audit
    row, so amounts are stringified on the way in.
    """
    call(client, scenario(), principal(Persona.MANAGER), url=url, ta_da="100")

    assert len(audit.events) == 1
    event = audit.events[0]
    assert event.action == action
    assert event.actor_id == STAFF_USER_ID
    assert event.actor_persona is Persona.MANAGER
    assert event.after is not None
    assert event.after["net"] == "14035"
    assert event.after["artifact_state"] == ArtifactState.DRAFT.value
    assert event.after["is_blocked"] is True
    assert not any(isinstance(value, float) for value in event.after.values())


def test_a_preview_writes_no_audit_row(client, audit) -> None:
    """A pure read is not a state transition. §11 says every transition writes one,
    not every request."""
    call(client, scenario(), principal(Persona.MANAGER))
    assert audit.events == []


def test_no_payout_route_can_release_anything() -> None:
    """R4/R3, as a test so it fails loudly if someone adds one.

    Approval and release require an authenticated human session against the
    approval state machine. There is no endpoint here that sends, releases, marks
    paid, or transitions an artifact — and every route is a POST that computes.
    """
    forbidden = ("release", "send", "approve", "paid", "pay:", "submit", "email", "issue")
    paths = [route.path for route in payouts.router.routes]

    assert paths, "the payout router registered no routes at all"
    for path in paths:
        assert not any(word in path.lower() for word in forbidden), path


def test_main_registers_the_payout_router(monkeypatch) -> None:
    """`app/main.py` mounts it, alphabetically, as its docstring instructs."""
    from app.core.config import get_settings
    from app.main import create_app

    for key, value in {
        "SUPABASE_URL": "https://example.supabase.co",
        "SUPABASE_ANON_KEY": "anon",
        "SUPABASE_SERVICE_ROLE_KEY": "service",
        "DATABASE_URL": "postgresql://postgres:pw@db.example.com:5432/postgres",
    }.items():
        monkeypatch.setenv(key, value)
    get_settings.cache_clear()
    try:
        # Asserted against the OpenAPI schema, NOT `app.routes`.
        #
        # `app.routes` is FastAPI's internal structure and its shape is not
        # stable: since 0.121, `include_router()` returns an `_IncludedRouter`
        # wrapper that holds the child routes instead of flattening them into
        # the parent list. Walking `app.routes` therefore found every mounted
        # endpoint on 0.136 and *none* of them on 0.141 — while the application
        # served both identically. `pyproject.toml` pins only `fastapi>=0.115`,
        # so both resolutions are legal and the old test failed on the newer one
        # for no defect.
        #
        # The schema is the public contract: it is what a client sees and what
        # the frontend generates against. Asserting on it tests the thing that
        # actually matters and cannot be broken by an internal refactor upstream.
        paths = set(create_app().openapi()["paths"])
    finally:
        get_settings.cache_clear()

    assert ENDPOINTS[0] in paths
    assert set(ENDPOINTS) <= paths
