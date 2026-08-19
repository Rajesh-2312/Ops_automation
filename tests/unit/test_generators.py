"""Sheet and invoice generator tests. CLAUDE.md §11 and §12.

Two things are being defended here and they are different:

1. **The column orders are an output contract.** Finance has VLOOKUPs written
   against these headers. A test that only checked "21 columns" would pass while
   swapping two of them, so the assertions are on the exact tuple, position by
   position, including the legacy misspelling.

2. **The §6 fixtures must reconcile to the rupee *in the file*.** `test_engine.py`
   already proves the engine computes them. That is not the same claim: a
   generator that read `gross` into the `Net Pay` column would leave every engine
   test green and pay every trainer the wrong amount. These tests re-assert both
   fixtures by reading cells back out of a written workbook.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.domain.enums import ProgramType, RateBasis
from app.domain.money import money
from app.domain.payout import PayoutInput
from app.services.remuneration.engine import compute_payout
from app.services.remuneration.generators import (
    INVOICE_COLUMNS,
    LEGACY_MISSPELLINGS,
    REMUNERATION_COLUMNS,
    PayoutRecord,
    TrainerIdentity,
    invoice_row,
    remuneration_columns,
    remuneration_row,
    write_invoice_sheet,
    write_remuneration_sheet,
)

# --- fixtures from CLAUDE.md §6 ----------------------------------------------


def _vema() -> PayoutRecord:
    """bCAP, Rs 80,000/mo, 26-31 Jul 2026, TA&DA Rs 100.

    Expected: Earned 15,484 - Gross 15,584 - TDS 1,548 - Net 14,035.
    """
    payout = PayoutInput(
        program_type=ProgramType.BCAP,
        rate_basis=RateBasis.PER_MONTH,
        rate=money(80_000),
        payable_days=Decimal(6),
        days_in_month=31,
        period_start=date(2026, 7, 26),
        period_end=date(2026, 7, 31),
        ta_da=money(100),
    )
    return PayoutRecord(
        payout=payout,
        result=compute_payout(payout),
        identity=TrainerIdentity(
            pan="BRMPV1234A",
            name="VEMA PRUDHVI SAI",
            bank_account_number="1234567890",
            ifsc="SBIN0001234",
            bank_name="State Bank of India",
            branch="Guntur",
            account_name="VEMA PRUDHVI SAI",
            email="vema@example.com",
        ),
        payout_month=date(2026, 7, 1),
        period_start=date(2026, 7, 26),
        period_end=date(2026, 7, 31),
        invoice_number="BRMP/26-27/JUL1",
        college_tags="MALINENI",
        college_name="Malineni Lakshmaiah Women's Engineering College",
        program_name="bCAP",
        invoice_date=date(2026, 8, 1),
    )


def _bushily() -> PayoutRecord:
    """bCAP, Rs 65,000/mo, full Jul 2026. Expected: Earned 65,000 - TDS 6,500 - Net 58,500.

    This is the row the legacy invoice sheet stores as `65000.00000000001`
    (findings §2). The workbook assertions below are what prove we do not.
    """
    payout = PayoutInput(
        program_type=ProgramType.BCAP,
        rate_basis=RateBasis.PER_MONTH,
        rate=money(65_000),
        payable_days=Decimal(31),
        days_in_month=31,
        period_start=date(2026, 7, 1),
        period_end=date(2026, 7, 31),
    )
    return PayoutRecord(
        payout=payout,
        result=compute_payout(payout),
        identity=TrainerIdentity(
            pan="BCDPK5678B",
            name="Bushily Kondala Rao",
            bank_account_number="9876543210",
            ifsc="HDFC0004321",
            bank_name="HDFC Bank",
            account_name="Bushily Kondala Rao",
        ),
        payout_month=date(2026, 7, 1),
        period_start=date(2026, 7, 1),
        period_end=date(2026, 7, 31),
        invoice_number="BCDP/26-27/JUL1",
        college_tags="MALINENI",
        program_name="bCAP",
    )


def _crt() -> PayoutRecord:
    """CRT, Rs 3,500/day x 6 days - the per-day path (findings §4)."""
    payout = PayoutInput(
        program_type=ProgramType.CRT,
        rate_basis=RateBasis.PER_DAY,
        rate=money(3_500),
        payable_days=Decimal(6),
        days_in_month=31,
        period_start=date(2026, 7, 13),
        period_end=date(2026, 7, 18),
    )
    return PayoutRecord(
        payout=payout,
        result=compute_payout(payout),
        identity=TrainerIdentity(pan="AFBPY9012C", name="Vijaya Raghava"),
        payout_month=date(2026, 7, 1),
        period_start=date(2026, 7, 13),
        period_end=date(2026, 7, 18),
        invoice_number="AFBP/26-27/JUL1",
    )


def _cells(workbook_bytes: BytesIO) -> list[tuple[object, ...]]:
    workbook = load_workbook(workbook_bytes)
    return [tuple(row) for row in workbook.active.iter_rows(values_only=True)]


# --- column order: the output contract ---------------------------------------


def test_remuneration_sheet_has_the_21_legacy_columns_in_order() -> None:
    assert remuneration_columns(date(2026, 7, 1)) == (
        "Sl. No.",
        "PAN",
        "College Tags",
        "Name",
        "Commercials",
        "Unit",
        "No.of days per month",
        "No. of days to be paid in JUL",
        "Commercials per day",
        "Earned",
        "TA & DA",
        "Accomodation",
        "Travel reimb",
        "Gross",
        "Deductions",
        "TDS",
        "Net Pay",
        "Bank AC no.",
        "IFSC",
        "Name of the Bank",
        "Invoice Number",
    )


def test_invoice_sheet_has_the_34_legacy_columns_in_order() -> None:
    assert INVOICE_COLUMNS == (
        "date",
        "Name",
        "PAN No.",
        "Expense for the month",
        "account_name",
        "College name",
        "Commercials(per month- Actual)",
        "Unit",
        "Attendance ( no.of worked days)",
        "Commercials per day",
        "Earned per month",
        "TA & DA",
        "Accomodation Allowance",
        "Travel reimb",
        "Gross",
        "Deductions",
        "TDS",
        "Net Pay",
        "Total Pay",
        "Bank AC no.",
        "IFSC",
        "Name of Bank",
        "Branch",
        "Invoice Number",
        "Address",
        "Contact Number",
        "Program Name",
        "From",
        "To",
        "Amount in Words",
        "Trainer Mail ID",
        "AM Mail ID",
        "HR Mail ID",
        "SMO Mail ID",
    )


def test_legacy_misspelling_is_preserved_in_both_sheets() -> None:
    """§11: people trust the format they read. Fixing the spelling breaks lookups."""
    assert "Accomodation" in REMUNERATION_COLUMNS
    assert "Accomodation Allowance" in INVOICE_COLUMNS
    assert "Accommodation" not in REMUNERATION_COLUMNS
    assert not any(c.startswith("Accommodation") for c in INVOICE_COLUMNS)
    assert {"Accomodation"} == LEGACY_MISSPELLINGS


@pytest.mark.parametrize(
    ("month", "expected"),
    [(date(2026, 7, 1), "JUL"), (date(2026, 1, 31), "JAN"), (date(2026, 12, 5), "DEC")],
)
def test_column_8_is_parameterised_by_payout_month(month: date, expected: str) -> None:
    assert remuneration_columns(month)[7] == f"No. of days to be paid in {expected}"


def test_row_width_matches_header_width() -> None:
    """A row that is one cell short silently shifts every column after it."""
    record = _vema()
    assert len(remuneration_row(record, serial=1)) == len(REMUNERATION_COLUMNS) == 21
    assert len(invoice_row(record)) == len(INVOICE_COLUMNS) == 34


# --- the fixtures, reconciled through the workbook ---------------------------


def test_vema_fixture_reconciles_in_the_written_remuneration_sheet() -> None:
    """CLAUDE.md §6: Earned 15,484 - Gross 15,584 - TDS 1,548 - Net 14,035."""
    stream = BytesIO()
    write_remuneration_sheet([_vema()], stream, payout_month=date(2026, 7, 1))

    header, row = _cells(stream)
    cell = dict(zip(header, row, strict=True))

    assert cell["PAN"] == "BRMPV1234A"
    assert cell["Commercials"] == Decimal(80_000)
    assert cell["Unit"] == "per month"
    assert cell["No.of days per month"] == 31
    assert cell["No. of days to be paid in JUL"] == Decimal(6)
    assert round(cell["Earned"]) == 15_484
    assert cell["TA & DA"] == Decimal(100)
    assert round(cell["Gross"]) == 15_584
    assert round(cell["TDS"]) == 1_548
    assert cell["Net Pay"] == Decimal(14_035)
    assert cell["Invoice Number"] == "BRMP/26-27/JUL1"


def test_bushily_fixture_writes_exact_65000_not_the_float_artifact() -> None:
    """findings §2: the legacy sheet stores 65000.00000000001. Ours must not.

    The equality below is exact and deliberate — `Decimal("65000.00000000001")`
    would fail it, which is the entire point of R7 and of putting `Decimal` into
    the cell rather than `float`.
    """
    stream = BytesIO()
    write_invoice_sheet([_bushily()], stream)

    header, row = _cells(stream)
    cell = dict(zip(header, row, strict=True))

    assert cell["Earned per month"] == Decimal(65_000)
    assert cell["Gross"] == Decimal(65_000)
    assert cell["TDS"] == Decimal(6_500)
    assert cell["Net Pay"] == Decimal(58_500)
    assert cell["Total Pay"] == Decimal(58_500)
    assert cell["Amount in Words"] == "Fifty Eight Thousand Five Hundred Rupees Only"


def test_amount_in_words_tracks_net_pay_and_is_never_the_excel_error() -> None:
    """§6: the legacy sheet renders #NAME? from a missing macro. Not reproduced."""
    row = dict(zip(INVOICE_COLUMNS, invoice_row(_vema()), strict=True))
    assert row["Amount in Words"] == "Fourteen Thousand and Thirty Five Rupees Only"
    assert row["Net Pay"] == Decimal(14_035)
    assert "#NAME?" not in str(row["Amount in Words"])


def test_crt_row_labels_the_per_day_unit_and_prints_the_day_rate() -> None:
    row = dict(zip(REMUNERATION_COLUMNS, remuneration_row(_crt(), serial=1), strict=True))
    assert row["Unit"] == "per day"
    assert row["Commercials"] == Decimal(3_500)
    # On the per-day path the display rate IS the contracted rate, so unlike
    # bCAP these two columns agree and `rate x days` reconciles to `Earned`.
    assert row["Commercials per day"] == Decimal(3_500)
    assert row["Earned"] == Decimal(21_000)


# --- the deliberately empty columns ------------------------------------------


def test_open_questions_write_empty_cells_rather_than_invented_values() -> None:
    """§14 Q2 and findings §1. An invented value is indistinguishable from a real one."""
    row = dict(zip(INVOICE_COLUMNS, invoice_row(_bushily()), strict=True))
    assert row["Expense for the month"] is None
    assert row["AM Mail ID"] is None


def test_supplied_open_question_values_are_still_written() -> None:
    """Empty by default, not hardcoded empty: a caller who knows may fill them."""
    record = _bushily()
    filled = PayoutRecord(
        payout=record.payout,
        result=record.result,
        identity=record.identity,
        payout_month=record.payout_month,
        period_start=record.period_start,
        period_end=record.period_end,
        invoice_number=record.invoice_number,
        am_mail_id="am@bytexl.in",
        expense_for_the_month=date(2026, 7, 26),
    )
    row = dict(zip(INVOICE_COLUMNS, invoice_row(filled), strict=True))
    assert row["AM Mail ID"] == "am@bytexl.in"
    assert row["Expense for the month"] == date(2026, 7, 26)


# --- guards ------------------------------------------------------------------


def test_mismatched_input_and_result_are_rejected() -> None:
    """Pairing one trainer's result with another's input reconciles column-wise
    and is entirely wrong. It must not be constructible."""
    vema = _vema()
    with pytest.raises(ValueError, match="were not computed|did not|disagree"):
        PayoutRecord(
            payout=_bushily().payout,
            result=vema.result,
            identity=vema.identity,
            payout_month=date(2026, 7, 1),
            period_start=date(2026, 7, 26),
            period_end=date(2026, 7, 31),
            invoice_number="BRMP/26-27/JUL1",
        )


def test_mixing_months_in_one_remuneration_sheet_is_rejected() -> None:
    """Column 8 names a single month; a June row under a JUL header is mislabelled."""
    june = _crt()
    august = PayoutRecord(
        payout=june.payout,
        result=june.result,
        identity=june.identity,
        payout_month=date(2026, 8, 1),
        period_start=june.period_start,
        period_end=june.period_end,
        invoice_number="AFBP/26-27/AUG1",
    )
    with pytest.raises(ValueError, match="another month"):
        write_remuneration_sheet([june, august], BytesIO(), payout_month=date(2026, 7, 1))


def test_pan_is_required_because_it_seeds_the_invoice_number() -> None:
    with pytest.raises(ValueError, match="pan is required"):
        TrainerIdentity(pan="   ", name="No PAN")


def test_empty_sheet_still_writes_a_correct_header() -> None:
    stream = BytesIO()
    write_remuneration_sheet([], stream, payout_month=date(2026, 3, 1))
    (header,) = _cells(stream)
    assert header[7] == "No. of days to be paid in MAR"


def test_serial_numbers_are_sequential_across_rows() -> None:
    stream = BytesIO()
    write_remuneration_sheet([_vema(), _bushily()], stream, payout_month=date(2026, 7, 1))
    _, first, second = _cells(stream)
    assert (first[0], second[0]) == (1, 2)


def test_generators_contain_no_arithmetic() -> None:
    """R2: every rupee is computed in engine.py and nowhere else.

    A crude but honest check — the module source must contain no arithmetic
    operator outside its prose. It is the cheapest way to make "this module
    computes nothing" a property the build enforces rather than a promise in a
    docstring.
    """
    import ast
    import inspect

    from app.services.remuneration import generators

    tree = ast.parse(inspect.getsource(generators))
    arithmetic = [
        node
        for node in ast.walk(tree)
        if isinstance(node, ast.BinOp)
        and isinstance(node.op, (ast.Mult, ast.Div, ast.Sub, ast.FloorDiv, ast.Mod))
    ]
    assert (
        arithmetic == []
    ), "generators.py performs arithmetic; money is computed in engine.py only (R2)"
