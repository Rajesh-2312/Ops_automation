"""Sheet and invoice generators. CLAUDE.md §11 — the column orders are contracts.

This module turns a computed payout into the two legacy outputs Finance already
reads: the 21-column remuneration sheet and the 34-column invoice sheet. Both
column orders are transcribed from the sample files in §15 and recorded in
`docs/legacy-sheet-findings.md`. §11: "Sheet outputs preserve legacy column
order. People trust the format they read."

**This module computes nothing.** R2 puts every rupee of arithmetic in
`engine.py`; a generator that derived even one figure would be a second place for
money to be wrong. Every numeric cell is read straight off the `PayoutInput` that
went in or the `PayoutResult` that came out, and the one derived string
(`Amount in Words`) is rendered from `net` by `app.domain.money.amount_in_words`.
There is no arithmetic operator anywhere in this file. If you need one, the value
belongs in `PayoutResult`.

**Why a record carries both input and result.** The sheets print three
reimbursement columns — `TA & DA`, `Accomodation`, `Travel reimb` — but
`PayoutResult` deliberately collapses them into a single `reimbursements` total,
and it carries `rate_per_day` (display, quantised) without the contracted `rate`.
Neither is recoverable after the fact: reconstructing a monthly rate as
`rate_per_day * days_in_month` re-multiplies a rounded value and yields 79,999.99
for an 80,000 engagement, which is precisely the R6 violation the engine exists to
avoid. So the record holds the `PayoutInput` too and reads those cells from it.
Pass-through, never re-derivation.

Three details that look like bugs and are not:

1. **`Accomodation` is misspelled** in both legacy headers. Preserved deliberately
   — see `LEGACY_MISSPELLINGS`. Correcting it silently breaks every downstream
   VLOOKUP Finance has written against the file.
2. **`Decimal` goes into the cell, not `float`.** openpyxl carries `Decimal` in
   its `NUMERIC_TYPES`, so the exact value is written. This is the whole reason
   the output does not reproduce the `65000.00000000001` in the shipped legacy
   invoice sheet (findings §2). Never call `float()` on the way to a cell.
3. **Two columns are deliberately left empty**: `Expense for the month` (§14 Q2 —
   semantics unknown) and `AM Mail ID` (findings §1 — "AM" is unidentified).
   §14 says carry the open questions, do not invent answers. An invented value
   here would be indistinguishable from a real one in a month.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from typing import IO, Final

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.enums import RateBasis
from app.domain.money import amount_in_words
from app.domain.payout import PayoutInput, PayoutResult
from app.services.remuneration.invoice_no import month_token

__all__ = [
    "INVOICE_COLUMNS",
    "LEGACY_MISSPELLINGS",
    "REMUNERATION_COLUMNS",
    "PayoutRecord",
    "TrainerIdentity",
    "invoice_columns",
    "invoice_row",
    "remuneration_columns",
    "remuneration_row",
    "write_invoice_sheet",
    "write_remuneration_sheet",
]


#: Headers the legacy files spell wrong. Kept in one greppable place so a future
#: reader finds the reason instead of "fixing" it. Asserted by a test.
LEGACY_MISSPELLINGS: Final[frozenset[str]] = frozenset({"Accomodation"})

#: Unit labels exactly as the legacy sheets write them.
_UNIT_LABEL: Final[dict[RateBasis, str]] = {
    RateBasis.PER_MONTH: "per month",
    RateBasis.PER_DAY: "per day",
}


# --- what the sheet needs that the engine never sees -------------------------


@dataclass(frozen=True, slots=True)
class TrainerIdentity:
    """Identity and payment rails: printed, never computed with.

    PAN is required and everything else is optional. §6 makes PAN the identity
    key ("never match trainers by name string"), and it is the one field whose
    absence must fail loudly rather than produce a blank cell — the invoice
    number is seeded from it.

    A missing optional field writes an empty cell. That is intentional: the §7
    gates already block a payout with no bank account or IFSC, so a blank here
    means the caller is previewing something that cannot be released, and it
    should look obviously incomplete rather than plausible.
    """

    pan: str
    name: str

    bank_account_number: str | None = None
    ifsc: str | None = None
    bank_name: str | None = None
    branch: str | None = None

    #: `account_name` on the invoice sheet — the beneficiary name on the bank
    #: account, which is not always the trainer's display name.
    account_name: str | None = None

    address: str | None = None
    contact_number: str | None = None
    email: str | None = None

    def __post_init__(self) -> None:
        if not self.pan.strip():
            raise ValueError(
                "TrainerIdentity.pan is required — §6 makes PAN the identity key "
                "and the invoice number is seeded from it"
            )


@dataclass(frozen=True, slots=True)
class PayoutRecord:
    """One payout — its input, its result, and everything needed to print it.

    Composition rather than a flat bag: `payout` and `result` are the engine's
    two halves and are the only source of a number in either sheet. Nothing in
    this class may shadow a field of either.
    """

    #: The structured input the engine was given. Source of the contracted rate
    #: and the three separate reimbursement columns — see the module docstring.
    payout: PayoutInput
    result: PayoutResult
    identity: TrainerIdentity

    #: The month being paid for. Drives the `No. of days to be paid in <MON>`
    #: header and the invoice number's month token. Any day within the month.
    payout_month: date

    period_start: date
    period_end: date

    #: Generated by `build_invoice_number()`, never trainer-supplied — findings
    #: §6: roughly one in five legacy numbers is malformed.
    invoice_number: str

    college_name: str | None = None
    #: `College Tags` on the remuneration sheet — the legacy short code.
    college_tags: str | None = None
    program_name: str | None = None

    #: The invoice's own `date` column. Explicit rather than defaulted to today:
    #: a regenerated invoice must reproduce the original date, and a generator
    #: that reaches for `date.today()` cannot.
    invoice_date: date | None = None

    # --- recipient routing, invoice sheet only ------------------------------
    am_mail_id: str | None = None
    """Unfilled by default. Findings §1: "AM" is an unidentified recipient.

    Present as a field so a caller who *knows* can supply it, absent from every
    default so the system never guesses who receives a trainer's invoice.
    """

    hr_mail_id: str | None = None
    smo_mail_id: str | None = None

    #: §14 Q2 — the legacy `Expense for the month` field holds `2026-07-26`
    #: against a 01–31 July period and its semantics are unknown. Left None so
    #: the column writes empty. Do not default it to the period end "because it
    #: looks like one".
    expense_for_the_month: date | None = None

    def __post_init__(self) -> None:
        """Reject an input/result pair that did not come from the same run.

        Cheap, and it catches the one mistake this class makes possible: pairing
        a result with somebody else's input, which would print one trainer's
        rate against another's earnings and reconcile perfectly at the column
        level while being entirely wrong.
        """
        mismatches = [
            name
            for name, left, right in (
                ("rate_basis", self.payout.rate_basis, self.result.rate_basis),
                ("payable_days", self.payout.payable_days, self.result.payable_days),
                ("days_in_month", self.payout.days_in_month, self.result.days_in_month),
                ("deductions", self.payout.deductions, self.result.deductions),
                ("reimbursements", self.payout.reimbursements, self.result.reimbursements),
            )
            if left != right
        ]
        if mismatches:
            raise ValueError(
                f"PayoutInput and PayoutResult disagree on {', '.join(mismatches)} — "
                "this result was not computed from this input"
            )
        if self.period_end < self.period_start:
            raise ValueError(
                f"period_end {self.period_end} precedes period_start {self.period_start}"
            )


# --- remuneration sheet: 21 columns ------------------------------------------
# docs/legacy-sheet-findings.md §1. Column 8 is month-parameterised, which is why
# `remuneration_columns()` exists and this tuple is its unbound template.

REMUNERATION_COLUMNS: Final[tuple[str, ...]] = (
    "Sl. No.",
    "PAN",
    "College Tags",
    "Name",
    "Commercials",
    "Unit",
    "No.of days per month",
    "No. of days to be paid in {month}",
    "Commercials per day",
    "Earned",
    "TA & DA",
    "Accomodation",  # sic — LEGACY_MISSPELLINGS
    "Travel reimb",
    "Gross",
    "Deductions",
    "TDS",
    "Net Pay",
    "Bank AC no.",
    "IFSC",
    "Name of the Bank",
    "Invoice Number",
)


def remuneration_columns(payout_month: date) -> tuple[str, ...]:
    """The 21 headers with column 8 bound to the payout month, e.g. `...in JUL`."""
    token = month_token(payout_month)
    return tuple(c.format(month=token) for c in REMUNERATION_COLUMNS)


def remuneration_row(record: PayoutRecord, *, serial: int) -> tuple[object, ...]:
    """One row, in `REMUNERATION_COLUMNS` order.

    `serial` is the caller's 1-based `Sl. No.` — a position in a printed sheet,
    not an identity, which is why it is a parameter and not a record field.

    Note `Commercials` is the contracted rate and `Commercials per day` is the
    engine's display-only `rate_per_day`. On the per-month path our
    `Commercials per day` x `days` will NOT equal our `Earned`. The legacy sheet's
    two columns disagree with each other in the same row too (findings §3) — the
    difference is that ours disagree in the direction of the correct answer.
    """
    payout, result, identity = record.payout, record.result, record.identity

    return (
        serial,
        identity.pan,
        record.college_tags,
        identity.name,
        payout.rate,
        _UNIT_LABEL[result.rate_basis],
        result.days_in_month,
        result.payable_days,
        result.rate_per_day,
        result.earned,
        payout.ta_da,
        payout.accommodation,
        payout.travel_reimbursement,
        result.gross,
        result.deductions,
        result.tds,
        result.net,
        identity.bank_account_number,
        identity.ifsc,
        identity.bank_name,
        record.invoice_number,
    )


# --- invoice sheet: 34 columns -----------------------------------------------

INVOICE_COLUMNS: Final[tuple[str, ...]] = (
    "date",
    "Name",
    "PAN No.",
    "Expense for the month",
    "account_name",
    "College name",
    "Commercials(per month- Actual)",
    "Unit",
    "Attendance ( no.of worked days)",
    "Commercials per day",
    "Earned per month",
    "TA & DA",
    "Accomodation Allowance",  # sic — LEGACY_MISSPELLINGS
    "Travel reimb",
    "Gross",
    "Deductions",
    "TDS",
    "Net Pay",
    "Total Pay",
    "Bank AC no.",
    "IFSC",
    "Name of Bank",
    "Branch",
    "Invoice Number",
    "Address",
    "Contact Number",
    "Program Name",
    "From",
    "To",
    "Amount in Words",
    "Trainer Mail ID",
    "AM Mail ID",
    "HR Mail ID",
    "SMO Mail ID",
)

#: 1-based positions of the date columns in `INVOICE_COLUMNS`, for number
#: formatting. Derived rather than hardcoded so reordering the tuple cannot
#: silently format the wrong columns.
_INVOICE_DATE_COLUMNS: Final[tuple[int, ...]] = tuple(
    INVOICE_COLUMNS.index(name) + 1 for name in ("date", "Expense for the month", "From", "To")
)


def invoice_columns() -> tuple[str, ...]:
    """The 34 headers. A function for symmetry with `remuneration_columns()`."""
    return INVOICE_COLUMNS


def invoice_row(record: PayoutRecord) -> tuple[object, ...]:
    """One row, in `INVOICE_COLUMNS` order.

    `Total Pay` is emitted as `net`. The legacy sheet carries both `Net Pay` and
    `Total Pay` and, in the one row we can reconcile (Bushily Kondala Rao), they
    hold the same figure — so this reproduces observed behaviour rather than a
    documented rule. If a real sheet ever shows the two diverging, the difference
    is a computed quantity and belongs in `PayoutResult`, not here.

    `Amount in Words` is rendered from `net`, the rounded whole-rupee figure, so
    the words and the `Net Pay` cell can never disagree. §6: the legacy sheet
    renders `#NAME?` from a missing Excel macro — not reproduced.
    """
    payout, result, identity = record.payout, record.result, record.identity

    return (
        record.invoice_date,
        identity.name,
        identity.pan,
        record.expense_for_the_month,  # §14 Q2 — empty by default, on purpose
        identity.account_name,
        record.college_name,
        payout.rate,
        _UNIT_LABEL[result.rate_basis],
        result.payable_days,
        result.rate_per_day,
        result.earned,
        payout.ta_da,
        payout.accommodation,
        payout.travel_reimbursement,
        result.gross,
        result.deductions,
        result.tds,
        result.net,
        result.net,  # Total Pay — see docstring
        identity.bank_account_number,
        identity.ifsc,
        identity.bank_name,
        identity.branch,
        record.invoice_number,
        identity.address,
        identity.contact_number,
        record.program_name,
        record.period_start,
        record.period_end,
        amount_in_words(result.net),
        identity.email,
        record.am_mail_id,  # findings §1 — unidentified recipient, empty by default
        record.hr_mail_id,
        record.smo_mail_id,
    )


# --- workbook writers --------------------------------------------------------


def write_remuneration_sheet(
    records: list[PayoutRecord],
    destination: str | IO[bytes],
    *,
    payout_month: date,
    sheet_title: str | None = None,
) -> Workbook:
    """Write the 21-column remuneration sheet and return the workbook.

    `payout_month` is a parameter rather than being read from the first record
    because an empty sheet still needs a correct month-parameterised header, and
    because rows straddling two months are a caller error that must not be
    silently resolved by whichever record happened to sort first.
    """
    _assert_single_month(records, payout_month)

    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = sheet_title or f"Remuneration {month_token(payout_month)}"

    sheet.append(list(remuneration_columns(payout_month)))
    for serial, record in enumerate(records, start=1):
        sheet.append(list(remuneration_row(record, serial=serial)))

    workbook.save(destination)
    return workbook


def write_invoice_sheet(
    records: list[PayoutRecord],
    destination: str | IO[bytes],
    *,
    sheet_title: str = "Invoices",
) -> Workbook:
    """Write the 34-column invoice sheet and return the workbook.

    No single-month assertion here: unlike the remuneration sheet, no invoice
    header is month-parameterised, and each row carries its own `From`/`To`.
    """
    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = sheet_title

    sheet.append(list(invoice_columns()))
    for record in records:
        sheet.append(list(invoice_row(record)))

    _format_dates(sheet, columns=_INVOICE_DATE_COLUMNS)
    workbook.save(destination)
    return workbook


# --- internals ---------------------------------------------------------------


def _active(workbook: Workbook) -> Worksheet:
    """The default sheet of a freshly created workbook.

    `Workbook.active` is `Worksheet | None` to the type checker because a loaded
    workbook can have no active sheet. A new one always does.
    """
    sheet = workbook.active
    if sheet is None:  # pragma: no cover — unreachable for a new Workbook
        raise RuntimeError("new workbook has no active sheet")
    return sheet


def _assert_single_month(records: list[PayoutRecord], payout_month: date) -> None:
    """Reject rows that do not belong to `payout_month`."""
    wrong = [
        r
        for r in records
        if (r.payout_month.year, r.payout_month.month) != (payout_month.year, payout_month.month)
    ]
    if wrong:
        pans = ", ".join(sorted(r.identity.pan for r in wrong))
        raise ValueError(
            f"records for another month in a {payout_month:%Y-%m} sheet: {pans}. "
            "Column 8 is month-parameterised; mixing months mislabels every row."
        )


def _format_dates(sheet: Worksheet, *, columns: tuple[int, ...]) -> None:
    """Apply a `DD-MM-YYYY` number format to the given 1-based date columns."""
    for column in columns:
        for row in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
            for cell in row:
                cell.number_format = "DD-MM-YYYY"
